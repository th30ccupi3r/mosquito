from io import BytesIO
import re

from openpyxl import load_workbook
from .models import Threat

HEADER_ALIASES = {
    "Component (TAM)": ["Component (TAM)"],
    "Threat category": ["Threat category"],
    "Questions": ["Questions"],
    "Typical threats": ["Typical threats"],
    "Mitigated (y/n)": ["Mitigated (y/n)"],
    "Attack path": ["Attack path"],
    "Impact (see matrix)": ["Impact (see matrix)"],
    "Easiness of attack (see matrix)": ["Easiness of attack (see matrix)"],
    "Risk severity (see matrix)": ["Risk severity (see matrix)"],
    "Comment / Follow-up": ["Comment / Follow-up", "Comment/Follow up"],
    "Link to mitigation (Jira, GitHub, ...)": [
        "Link to mitigation (Jira, GitHub, ...)",
        "Jira to migitation (Jira, Github, ...)",
    ],
    "Threat info": ["Threat info"],
    "Worst Case": ["Worst Case"],
}


def _replace_component_for_export(question: str, enabled: bool) -> str:
    if not enabled:
        return question

    pattern = re.compile(r"(?<!new )\bcomponent\b", re.IGNORECASE)

    def repl(match: re.Match[str]) -> str:
        word = match.group(0)
        if word.isupper():
            return "NEW COMPONENT"
        if word[0].isupper():
            return "New component"
        return "new component"

    return pattern.sub(repl, question)


def append_threats_to_workbook(
    file_bytes: bytes,
    threats: list[Threat],
    replace_component_with_new_component: bool = False,
) -> bytes:

    wb = load_workbook(filename=BytesIO(file_bytes))

    if "Threats" not in wb.sheetnames:
        raise ValueError("Worksheet 'Threats' not found")

    ws = wb["Threats"]

    headers = [cell.value for cell in ws[1]]
    raw_header_positions = {
        str(header).strip(): index
        for index, header in enumerate(headers, start=1)
        if isinstance(header, str) and header.strip()
    }
    header_positions: dict[str, int] = {}
    missing: list[str] = []

    for canonical_header, aliases in HEADER_ALIASES.items():
        matched_column = next(
            (raw_header_positions[alias] for alias in aliases if alias in raw_header_positions),
            None,
        )
        if matched_column is None:
            missing.append(canonical_header)
            continue
        header_positions[canonical_header] = matched_column

    if missing:
        raise ValueError("Missing headers: " + ", ".join(missing))

    for threat in threats:
        row_index = ws.max_row + 1
        row_values = {
            "Component (TAM)": "",
            "Threat category": threat.threat_category,
            "Questions": _replace_component_for_export(
                threat.question, replace_component_with_new_component
            ),
            "Typical threats": threat.typical_threat,
            "Mitigated (y/n)": "",
            "Attack path": "",
            "Impact (see matrix)": threat.impact,
            "Easiness of attack (see matrix)": threat.easiness_of_attack,
            "Risk severity (see matrix)": "",
            "Comment / Follow-up": "",
            "Link to mitigation (Jira, GitHub, ...)": "",
            "Threat info": "",
            "Worst Case": "",
        }

        for header, value in row_values.items():
            ws.cell(row=row_index, column=header_positions[header], value=value)

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output.read()
